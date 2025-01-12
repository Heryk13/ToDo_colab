"""Rotas de cada aplicação."""

from datetime import datetime

import openpyxl
from fastapi import APIRouter

from api.homepage.schemas import CreateList, ResponseList

router = APIRouter()

@router.get("/list_task")
def list_task() ->list[ResponseList] :
    """Creia uma lista de task predefinida.

    Returns:
        list[dict]: uma lista de tasks, onde cada task é representada por um
         dict.
        contendo uma chave:
            - "date" (datetime): o horario em que a task foi criada.
            - "title" (str): o titulo da task.
            - "description" (str): a descrição da task.
    """
    #carregando arquivo
    book = openpyxl.load_workbook("planilha.xlsx")
    #selecionando uma pagina
    data_page = book["data"]
    #print dos valores
    data_list = []
    for rows in data_page.iter_rows(min_row=2):
        date = datetime.strptime(rows[1].value, "%Y/%m/%d %H:%M:%S")
        title = rows[2].value
        description = rows[3].value
        data_row = {"date":date, "title":title, "description":description}
        data_list.append(data_row)

        #print(f'{rows[0].value}, {rows[1].value}, {rows[2].value}')
    return data_list

@router.post("/create_task")
def create_task(create_list : CreateList):
    """Cria uma nova task.

    Returns:
        list[dict]: .
    """
     #carregando arquivo
    book = openpyxl.load_workbook("planilha.xlsx")
    #selecionando uma pagina
    data_page = book["data"]
    #ler numeros de linhas
    num_rows = len(data_page["A"])
    #criar id novo
    create_id = data_page[f"A{num_rows}"].value + 1
    #adicionar uma nova linha
    create_data = [
        create_id,
        create_list.date.strftime("%Y/%m/%d %H:%M:%S"),
        create_list.title,
        create_list.description,
    ]

    data_page.append(create_data)

    book.save("planilha.xlsx")
    return create_list
    #dados.cell(row=1 colum=1) = bruno

@router.delete("/delete_task/{data_id}")
def delete_task(data_id: str):
    """Deleta uma task.

    Returns:
        list[dict]:.
    """
    #carregando arquivo
    book = openpyxl.load_workbook("planilha.xlsx")
    #selecionando uma pagina
    data_page = book.active
    #corresponder ao id do usuario na lista
    for index, _row in enumerate(data_page.iter_rows(1), 1):
        if data_page[f"A{index}"].value == int(data_id):
            data_page.delete_rows(index)
            book.save("planilha.xlsx")
            break

